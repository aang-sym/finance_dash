import csv
import json
import re
from datetime import date, datetime, timedelta
from datetime import datetime as dt_class
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from dateutil import parser
from flask import Flask, jsonify, redirect, request, send_file
from openpyxl import load_workbook

from sync import CONFIG_PATH, DATA_DIR, discover_account_ids, load_config, sync_transactions


BASE_DIR = Path(__file__).resolve().parent
BILL_CYCLES_PATH = DATA_DIR / "bill_cycles.csv"
BILL_TYPES_PATH = DATA_DIR / "bill_types.csv"
NETWORTH_CSV = DATA_DIR / "networth.csv"
HOLDINGS_CSV = DATA_DIR / "holdings.csv"
EXCEL_PATH = DATA_DIR / "Net worth calculator.xlsx"
NETWORTH_FIELDS = ["date", "cash_aud", "investments_aud", "super_aud", "total_aud"]
HOLDINGS_FIELDS = [
    "ticker", "platform", "currency", "units",
    "cost_base_aud", "current_price_aud", "current_value_aud",
    "unrealised_gain_aud", "acquisition_date",
]
HOUSEMATE_NAMES = ["angus", "sean", "alex", "jarrod", "ryan"]
BILL_SLUGS = {"rent", "elec", "water", "internet", "gas"}
MONTHS = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
PERIOD_TAG_RE = re.compile(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)-(\d{4})$")
LABEL_DEFAULTS = {
    "rent": "Rent",
    "elec": "Electricity",
    "water": "Water",
    "internet": "Internet",
    "gas": "Gas",
}

app = Flask(__name__, static_folder=str(BASE_DIR), static_url_path="")


def read_csv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader)


def write_csv(path: Path, fieldnames: List[str], rows: List[Dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_float(value: Optional[str]) -> Optional[float]:
    if value in (None, ""):
        return None
    return float(value)


def parse_bool(value: Optional[str]) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def parse_date(value: str) -> date:
    return date.fromisoformat(value)


def cycle_tag_for(slug: str, due_date: str) -> Tuple[str, str, str]:
    due = parse_date(due_date)
    month = MONTHS[due.month - 1]
    year = str(due.year)
    return f"{slug}-{month}-{year}", month, year


def parse_cycle_tag(tag: str) -> Optional[Tuple[str, str, str]]:
    parts = [part.strip().lower() for part in (tag or "").split("-")]
    if len(parts) != 3:
        return None
    slug, month, year = parts
    if slug not in BILL_SLUGS or month not in MONTHS or not year.isdigit():
        return None
    return slug, month, year


def parse_period_tag(tag: str) -> Optional[Tuple[str, str]]:
    match = PERIOD_TAG_RE.match((tag or "").strip().lower())
    if not match:
        return None
    return match.group(1), match.group(2)


def expand_bill_cycle_tags(tags: set[str]) -> set[str]:
    expanded = set()
    slug_tags = {tag for tag in tags if tag in BILL_SLUGS}
    period_tags = [parse_period_tag(tag) for tag in tags]
    valid_periods = [period for period in period_tags if period]

    for tag in tags:
        parsed = parse_cycle_tag(tag)
        if parsed:
            slug, month, year = parsed
            expanded.add(f"{slug}-{month}-{year}")

    for slug in slug_tags:
        for month, year in valid_periods:
            expanded.add(f"{slug}-{month}-{year}")

    return expanded


def format_share(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    return round(value, 2)


def read_housemates() -> List[Dict[str, str]]:
    return read_csv(DATA_DIR / "housemates.csv")


def read_bill_types() -> Dict[str, Dict[str, str]]:
    rows = read_csv(BILL_TYPES_PATH)
    return {row["slug"]: row for row in rows if row.get("slug")}


def read_bills() -> List[Dict[str, str]]:
    return read_csv(DATA_DIR / "bills.csv")


def read_manual_overrides() -> List[Dict[str, str]]:
    return read_csv(DATA_DIR / "manual_overrides.csv")


def write_bill_cycles(rows: List[Dict[str, str]]) -> None:
    write_csv(
        BILL_CYCLES_PATH,
        [
            "cycle_id",
            "slug",
            "label",
            "provider",
            "month",
            "year",
            "total_due",
            "collected_amount",
            "forwarded_amount",
            "paid_housemates",
            "paid_count",
            "status",
            "latest_activity",
        ],
        rows,
    )


def get_status() -> Dict:
    if not CONFIG_PATH.exists():
        config = {
            "token": "",
            "account_ids": {"spending": "", "savings": "", "grow": "", "two_up": ""},
            "last_sync_spending": None,
            "last_sync_2up": None,
        }
    else:
        with CONFIG_PATH.open("r", encoding="utf-8") as handle:
            config = json.load(handle)
    return {
        "last_sync_spending": config.get("last_sync_spending"),
        "last_sync_2up": config.get("last_sync_2up"),
        "account_ids": config.get("account_ids", {}),
        "token_present": bool(config.get("token")),
    }


def get_budgets() -> Dict[str, float]:
    if not CONFIG_PATH.exists():
        return {}
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    return {k: float(v) for k, v in config.get("budgets", {}).items()}


def save_budgets(budgets: Dict[str, float]) -> None:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError("config.json not found")
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    config["budgets"] = budgets
    with CONFIG_PATH.open("w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2)
        handle.write("\n")


def is_incoming_beem(row: Dict[str, str]) -> bool:
    amount = parse_float(row.get("amount")) or 0.0
    description = (row.get("description") or "").lower()
    transfer_account_id = (row.get("transfer_account_id") or "").strip()
    return amount > 0 and ("beem" in description or not transfer_account_id)


def tag_set(tags: str) -> set[str]:
    return {tag.strip().lower() for tag in (tags or "").split(",") if tag.strip()}


def build_override_lookup() -> Dict[Tuple[str, str, str, str], Dict[str, str]]:
    lookup: Dict[Tuple[str, str, str, str], Dict[str, str]] = {}
    for row in read_manual_overrides():
        key = (
            (row.get("housemate") or "").lower(),
            (row.get("slug") or "").lower(),
            (row.get("month") or "").lower(),
            str(row.get("year") or ""),
        )
        lookup[key] = row
    return lookup


def share_for_housemate(bill: Dict[str, str], housemate: Dict[str, str], total_amount: Optional[float]) -> Optional[float]:
    split_type = (bill.get("split_type") or "").lower()
    if split_type == "fixed":
        return format_share(parse_float(housemate.get("rent_share")))
    if split_type == "equal" and total_amount is not None:
        return format_share(total_amount / len(HOUSEMATE_NAMES))
    return None


def total_due_for_bill(bill: Dict[str, str], housemates: List[Dict[str, str]]) -> Optional[float]:
    split_type = (bill.get("split_type") or "").lower()
    if split_type == "fixed":
        return round(sum(parse_float(housemate.get("rent_share")) or 0.0 for housemate in housemates), 2)
    return parse_float(bill.get("total_amount"))


def find_bill_template(slug: str) -> Optional[Dict[str, str]]:
    bills = [bill for bill in read_bills() if (bill.get("slug") or "").lower() == slug.lower()]
    if not bills:
        return None
    bills.sort(key=lambda bill: bill.get("due_date", ""))
    return bills[0]


def cycle_bill_model(slug: str, month: str, year: str) -> Dict[str, Optional[str]]:
    cycle_due_date = f"{year}-{MONTHS.index(month) + 1:02d}-01"
    matching = [
        bill for bill in read_bills()
        if (bill.get("slug") or "").lower() == slug.lower()
        and cycle_tag_for(bill["slug"], bill["due_date"])[0] == f"{slug}-{month}-{year}"
    ]
    template = matching[0] if matching else find_bill_template(slug)
    if template:
        model = dict(template)
        model["slug"] = slug
        model["label"] = template.get("label") or LABEL_DEFAULTS.get(slug, slug.title())
        if not matching:
            model["due_date"] = cycle_due_date
        return model

    return {
        "slug": slug,
        "label": LABEL_DEFAULTS.get(slug, slug.title()),
        "due_date": cycle_due_date,
        "recurrence": "",
        "split_type": "fixed" if slug == "rent" else "equal",
        "notes": "",
        "total_amount": "",
        "status": "pending",
    }


def infer_cycle_from_date(slug: str, created_at: str) -> Tuple[str, str, str]:
    dt = parse_datetime_or_date(created_at)
    month = MONTHS[dt.month - 1]
    year = str(dt.year)
    return f"{slug}-{month}-{year}", month, year


def match_bill_type_for_two_up_payment(row: Dict[str, str], bill_types: Dict[str, Dict[str, str]]) -> Optional[str]:
    description = (row.get("description") or "").strip().lower()
    message = (row.get("message") or "").strip().lower()
    raw_text = (row.get("raw_text") or "").strip().lower()
    haystack = " ".join(part for part in [description, message, raw_text] if part)

    for slug, bill_type in bill_types.items():
        provider = (bill_type.get("provider") or "").strip().lower()
        label = (bill_type.get("label") or "").strip().lower()
        if provider and provider in haystack:
            return slug
        if label and label in haystack:
            return slug
    return None


def compute_bill_statuses() -> List[Dict]:
    bills = read_bills()
    housemates = read_housemates()
    transactions = read_csv(DATA_DIR / "transactions_spending.csv")
    overrides = build_override_lookup()
    today = date.today()

    beem_transactions = [row for row in transactions if is_incoming_beem(row)]
    bill_payloads: List[Dict] = []

    for bill in bills:
        due_date = bill["due_date"]
        due = parse_date(due_date)
        total_amount = total_due_for_bill(bill, housemates)
        cycle_tag, month, year = cycle_tag_for(bill["slug"], due_date)
        housemate_rows = []
        total_collected = 0.0

        for housemate in housemates:
            name = (housemate.get("name") or "").lower()
            share = share_for_housemate(bill, housemate, total_amount)
            paid = False
            source = None

            for txn in beem_transactions:
                tags = tag_set(txn.get("tags", ""))
                expanded_cycle_tags = expand_bill_cycle_tags(tags)
                if name in tags and cycle_tag in expanded_cycle_tags:
                    paid = True
                    source = "tag"
                    break

            if not paid:
                override = overrides.get((name, bill["slug"].lower(), month, year))
                if override is not None:
                    paid = parse_bool(override.get("paid"))
                    source = "manual" if paid else None

            if paid and share is not None:
                total_collected += share

            housemate_rows.append(
                {
                    "name": name,
                    "share": share,
                    "paid": paid,
                    "source": source,
                }
            )

        outstanding = None if total_amount is None else round(max(total_amount - total_collected, 0.0), 2)
        derived_status = bill.get("status", "pending")
        paid_count = sum(1 for entry in housemate_rows if entry["paid"])
        if paid_count == len(housemate_rows) and housemate_rows:
            derived_status = "paid"
        elif paid_count > 0:
            derived_status = "partial"
        else:
            derived_status = "pending"

        if due <= today + timedelta(days=60) or derived_status != "paid":
            bill_payloads.append(
                {
                    "id": int(bill["id"]),
                    "slug": bill["slug"],
                    "label": bill["label"],
                    "total_amount": total_amount,
                    "due_date": due_date,
                    "recurrence": bill["recurrence"],
                    "split_type": bill["split_type"],
                    "notes": bill.get("notes", ""),
                    "status": derived_status,
                    "cycle_tag": cycle_tag,
                    "housemates": housemate_rows,
                    "total_collected": round(total_collected, 2),
                    "outstanding": outstanding,
                }
            )

    bill_payloads.sort(key=lambda entry: entry["due_date"])
    return bill_payloads


def compute_bill_history() -> List[Dict]:
    spending_rows = read_csv(DATA_DIR / "transactions_spending.csv")
    two_up_rows = read_csv(DATA_DIR / "transactions_2up.csv")
    status = get_status()
    two_up_id = status.get("account_ids", {}).get("two_up", "")
    housemates = read_housemates()
    bill_types = read_bill_types()
    history: Dict[str, Dict] = {}

    def ensure_entry(cycle_tag: str, slug: str, month: str, year: str, latest_activity: str, description: str = "") -> Dict:
        entry = history.setdefault(
            cycle_tag,
            {
                "cycle_id": cycle_tag,
                "cycle_tag": cycle_tag,
                "slug": slug,
                "label": bill_types.get(slug, {}).get("label") or LABEL_DEFAULTS.get(slug, slug.title()),
                "month": month,
                "year": int(year),
                "housemates_paid": set(),
                "forwarded_total": 0.0,
                "matched_transactions": 0,
                "latest_activity": latest_activity,
                "descriptions": set(),
                "seeded_from_bill_payment": False,
                "bill_payment_amount": None,
                "paid_date": None,
            },
        )
        if latest_activity and latest_activity > entry["latest_activity"]:
            entry["latest_activity"] = latest_activity
        if description:
            entry["descriptions"].add(description)
        return entry

    for row in two_up_rows:
        amount = parse_float(row.get("amount")) or 0.0
        if amount >= 0:
            continue
        tags = tag_set(row.get("tags", ""))
        created_at = row.get("settled_at") or row.get("created_at") or ""
        description = row.get("description", "") or ""

        cycle_tags = sorted(expand_bill_cycle_tags(tags))
        if cycle_tags:
            inferred_cycles = cycle_tags
        else:
            matched_slug = match_bill_type_for_two_up_payment(row, bill_types)
            if not matched_slug or not created_at:
                continue
            inferred_cycle, month, year = infer_cycle_from_date(matched_slug, created_at)
            cycle_tags = [inferred_cycle]
            inferred_cycles = cycle_tags

        for cycle_tag in inferred_cycles:
            parsed = parse_cycle_tag(cycle_tag)
            if not parsed:
                continue
            slug, month, year = parsed
            entry = ensure_entry(cycle_tag, slug, month, year, created_at, description)
            entry["matched_transactions"] += 1
            entry["seeded_from_bill_payment"] = True
            entry["bill_payment_amount"] = round(abs(amount), 2)
            if not entry["paid_date"] or created_at < entry["paid_date"]:
                entry["paid_date"] = created_at

    for row in spending_rows:
        tags = tag_set(row.get("tags", ""))
        cycle_tags = sorted(expand_bill_cycle_tags(tags))
        if not cycle_tags:
            continue

        amount = parse_float(row.get("amount")) or 0.0
        transfer_account_id = row.get("transfer_account_id", "")
        created_at = row.get("settled_at") or row.get("created_at") or ""
        description = row.get("description", "") or ""

        for cycle_tag in cycle_tags:
            slug, month, year = parse_cycle_tag(cycle_tag)  # type: ignore[misc]
            entry = ensure_entry(cycle_tag, slug, month, year, created_at, description)
            entry["matched_transactions"] += 1

            if is_incoming_beem(row):
                paid_names = tags.intersection(HOUSEMATE_NAMES)
                entry["housemates_paid"].update(paid_names)

            if transfer_account_id and transfer_account_id == two_up_id and amount < 0:
                entry["forwarded_total"] += abs(amount)

    overrides = build_override_lookup()
    payload = []
    cycle_csv_rows: List[Dict[str, str]] = []
    for entry in history.values():
        bill_model = cycle_bill_model(entry["slug"], entry["month"], str(entry["year"]))
        total_due = total_due_for_bill(bill_model, housemates)
        if entry.get("seeded_from_bill_payment") and entry.get("bill_payment_amount") is not None:
            # Always trust the actual payment amount over the current config
            total_due = float(entry["bill_payment_amount"])

        # Merge manual overrides into housemates_paid
        for housemate in housemates:
            name = (housemate.get("name") or "").lower()
            if name not in entry["housemates_paid"]:
                override = overrides.get((name, entry["slug"].lower(), entry["month"], str(entry["year"])))
                if override and parse_bool(override.get("paid")):
                    entry["housemates_paid"].add(name)

        collected_amount = 0.0
        for housemate in housemates:
            name = (housemate.get("name") or "").lower()
            if name in entry["housemates_paid"]:
                share = share_for_housemate(bill_model, housemate, total_due)
                if share is not None:
                    collected_amount += share

        if entry.get("seeded_from_bill_payment"):
            # Outbound 2Up payment to provider confirms bill was paid — no tagging needed
            status_value = "paid"
        elif entry["housemates_paid"] and len(entry["housemates_paid"]) == len(housemates):
            status_value = "paid"
        elif entry["housemates_paid"]:
            status_value = "partial"
        else:
            status_value = "pending"
        provider = bill_types.get(entry["slug"], {}).get("provider", "")

        payload.append(
            {
                "cycle_id": entry["cycle_id"],
                "cycle_tag": entry["cycle_tag"],
                "slug": entry["slug"],
                "label": bill_model.get("label") or entry["label"],
                "provider": provider,
                "month": entry["month"],
                "year": entry["year"],
                "housemates_paid": sorted(entry["housemates_paid"]),
                "housemates_paid_count": len(entry["housemates_paid"]),
                "incoming_total": round(collected_amount, 2),
                "forwarded_total": round(entry["forwarded_total"], 2),
                "matched_transactions": entry["matched_transactions"],
                "latest_activity": entry["latest_activity"],
                "paid_date": entry.get("paid_date"),
                "descriptions": sorted(entry["descriptions"]),
                "total_due": total_due,
                "status": status_value,
            }
        )
        cycle_csv_rows.append(
            {
                "cycle_id": entry["cycle_id"],
                "slug": entry["slug"],
                "label": bill_model.get("label") or entry["label"],
                "provider": provider,
                "month": entry["month"],
                "year": str(entry["year"]),
                "total_due": "" if total_due is None else f"{total_due:.2f}",
                "collected_amount": f"{collected_amount:.2f}",
                "forwarded_amount": f"{entry['forwarded_total']:.2f}",
                "paid_housemates": ",".join(sorted(entry["housemates_paid"])),
                "paid_count": str(len(entry["housemates_paid"])),
                "status": status_value,
                "latest_activity": entry["latest_activity"],
            }
        )

    payload.sort(
        key=lambda item: (
            item["year"],
            MONTHS.index(item["month"]),
            item["slug"],
        ),
        reverse=True,
    )
    cycle_csv_rows.sort(
        key=lambda item: (
            item["year"],
            MONTHS.index(item["month"]),
            item["slug"],
        ),
        reverse=True,
    )
    write_bill_cycles(cycle_csv_rows)
    return payload


def append_bill(payload: Dict) -> Dict:
    bills_path = DATA_DIR / "bills.csv"
    bills = read_bills()
    next_id = max((int(row["id"]) for row in bills), default=0) + 1
    row = {
        "id": str(next_id),
        "slug": payload["slug"],
        "label": payload["label"],
        "total_amount": "" if payload.get("total_amount") in (None, "") else f"{float(payload['total_amount']):.2f}",
        "due_date": payload["due_date"],
        "recurrence": payload["recurrence"],
        "split_type": payload["split_type"],
        "notes": payload.get("notes", ""),
        "status": "pending",
    }
    bills.append(row)
    write_csv(
        bills_path,
        ["id", "slug", "label", "total_amount", "due_date", "recurrence", "split_type", "notes", "status"],
        bills,
    )
    row["id"] = next_id
    row["total_amount"] = parse_float(row["total_amount"])
    return row


def upsert_override(bill_id: int, payload: Dict) -> Dict:
    bills = {int(row["id"]): row for row in read_bills()}
    if bill_id not in bills:
        raise KeyError(f"Bill {bill_id} not found")

    bill = bills[bill_id]
    cycle_tag, month, year = cycle_tag_for(bill["slug"], bill["due_date"])
    _ = cycle_tag
    housemate = (payload.get("housemate") or "").lower().strip()
    if housemate not in HOUSEMATE_NAMES:
        raise ValueError("Unknown housemate")

    paid = bool(payload.get("paid"))
    note = payload.get("note", "")
    overrides = read_manual_overrides()
    fieldnames = ["housemate", "slug", "month", "year", "paid", "note"]

    updated = False
    for row in overrides:
        if (
            row.get("housemate", "").lower() == housemate
            and row.get("slug", "").lower() == bill["slug"].lower()
            and row.get("month", "").lower() == month
            and str(row.get("year", "")) == year
        ):
            row["paid"] = "true" if paid else "false"
            row["note"] = note
            updated = True
            break

    if not updated:
        overrides.append(
            {
                "housemate": housemate,
                "slug": bill["slug"],
                "month": month,
                "year": year,
                "paid": "true" if paid else "false",
                "note": note,
            }
        )

    write_csv(DATA_DIR / "manual_overrides.csv", fieldnames, overrides)
    return {
        "housemate": housemate,
        "slug": bill["slug"],
        "month": month,
        "year": year,
        "paid": paid,
        "note": note,
    }


def parse_datetime_or_date(value: str) -> datetime:
    parsed = parser.isoparse(value)
    if parsed.tzinfo is None:
        return parsed
    return parsed.astimezone().replace(tzinfo=None)


def filter_spending_rows(rows: List[Dict[str, str]], since: Optional[str], until: Optional[str], category: Optional[str]) -> List[Dict]:
    status = get_status()
    account_ids = status.get("account_ids", {})
    two_up_id = account_ids.get("two_up", "")
    savings_id = account_ids.get("savings", "")
    grow_id = account_ids.get("grow", "")
    internal_ids = {tid for tid in (two_up_id, savings_id, grow_id) if tid}

    since_dt = parser.isoparse(since) if since else None
    until_dt = parser.isoparse(until) if until else None
    if until_dt and until_dt.hour == 0 and until_dt.minute == 0 and until_dt.second == 0:
        until_dt = until_dt + timedelta(days=1)

    filtered = []
    for row in rows:
        amount = parse_float(row.get("amount")) or 0.0
        transfer_account_id = row.get("transfer_account_id", "")

        # Skip all internal transfers (to/from savings, grow, 2up)
        if transfer_account_id in internal_ids:
            continue

        created_at = row.get("settled_at") or row.get("created_at")
        if not created_at:
            continue
        created_dt = parse_datetime_or_date(created_at)

        if since_dt and created_dt < since_dt.replace(tzinfo=None):
            continue
        if until_dt and created_dt >= until_dt.replace(tzinfo=None):
            continue
        if category and (row.get("category") or "").lower() != category.lower():
            continue

        filtered.append(
            {
                **row,
                "amount": round(amount, 2),
            }
        )

    filtered.sort(key=lambda item: item.get("settled_at") or item.get("created_at") or "", reverse=True)
    return filtered


def group_spending_by_period(rows: List[Dict], group_by: str) -> List[Dict]:
    buckets: Dict[str, float] = {}
    counts: Dict[str, int] = {}

    for row in rows:
        amount = float(row.get("amount", 0))
        if amount >= 0:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        dt = parse_datetime_or_date(dt_str)
        if group_by == "week":
            period = f"{dt.isocalendar()[0]}-W{dt.isocalendar()[1]:02d}"
        else:
            period = f"{dt.year}-{dt.month:02d}"
        buckets[period] = buckets.get(period, 0.0) + abs(amount)
        counts[period] = counts.get(period, 0) + 1

    return [
        {"period": period, "total_spend": round(buckets[period], 2), "transaction_count": counts[period]}
        for period in sorted(buckets.keys())
    ]


RECURRING_TRANSACTION_TYPES = {"Direct Debit", "Scheduled Transfer"}


def detect_recurring(rows: List[Dict[str, str]], exclusions: Optional[set] = None) -> List[Dict]:
    from statistics import median, stdev

    exclusions = exclusions or set()
    today = datetime.now().date()

    # Group transactions by description, tagging whether they are typed as recurring
    merchant_txns: Dict[str, List[Dict]] = {}
    for row in rows:
        amount = parse_float(row.get("amount")) or 0.0
        if amount >= 0:
            continue
        desc = (row.get("description") or "").strip()
        if not desc or desc in exclusions:
            continue
        transfer_id = row.get("transfer_account_id", "")
        if transfer_id:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        txn_type = (row.get("transaction_type") or "").strip()
        merchant_txns.setdefault(desc, []).append({
            "amount": abs(amount),
            "date": parse_datetime_or_date(dt_str),
            "typed_recurring": txn_type in RECURRING_TRANSACTION_TYPES,
        })

    recurring = []
    for desc, txns in merchant_txns.items():
        if len(txns) < 2:
            continue
        txns_sorted = sorted(txns, key=lambda t: t["date"])
        amounts = [t["amount"] for t in txns_sorted]
        dates = [t["date"] for t in txns_sorted]
        intervals = [(dates[i + 1] - dates[i]).days for i in range(len(dates) - 1)]
        med_interval = median(intervals)
        med_amount = median(amounts)
        if med_amount <= 0:
            continue

        # Primary signal: any transaction has a Direct Debit / Scheduled Transfer type
        typed = any(t["typed_recurring"] for t in txns)

        if typed:
            # Accept if interval looks roughly monthly (14–45 days) or annual (330–400)
            if not (14 <= med_interval <= 45 or 330 <= med_interval <= 400):
                continue
        else:
            # Fallback: strict statistical test (original behaviour)
            if not (20 <= med_interval <= 40):
                continue
            cv = (stdev(amounts) / med_amount) if len(amounts) > 1 else 0.0
            if cv > 0.15:
                continue

        monthly_cost = round(med_amount * (30.0 / med_interval), 2)
        last_date = dates[-1].date() if hasattr(dates[-1], "date") else dates[-1]
        next_date = last_date + timedelta(days=int(med_interval))

        # Skip entries where next expected date is more than one full interval in the past
        # (pattern likely stopped)
        if next_date < today - timedelta(days=int(med_interval)):
            continue

        recurring.append({
            "description": desc,
            "monthly_cost": monthly_cost,
            "median_amount": round(med_amount, 2),
            "interval_days": round(med_interval, 1),
            "occurrences": len(txns),
            "last_date": last_date.strftime("%Y-%m-%d"),
            "next_expected": next_date.strftime("%Y-%m-%d"),
        })

    return sorted(recurring, key=lambda r: r["monthly_cost"], reverse=True)


FRIVOLITY_WEIGHTS: Dict[str, float] = {
    "takeaway": 0.90,
    "restaurants-and-cafes": 0.75,
    "booze": 0.85,
    "pubs-and-bars": 0.85,
    "events-and-gigs": 0.80,
    "holidays-and-travel": 0.70,
    "hobbies": 0.70,
    "games-and-software": 0.80,
    "tv-and-music": 0.70,
    "lottery-and-gambling": 0.95,
    "clothing-and-accessories": 0.50,
    "hair-and-beauty": 0.50,
    "fitness-and-wellbeing": 0.30,
    "gifts-and-charity": 0.60,
    "technology": 0.60,
    "groceries": 0.10,
    "health-and-medical": 0.05,
    "rent-and-mortgage": 0.00,
    "utilities": 0.00,
    "internet": 0.05,
    "fuel": 0.10,
    "public-transport": 0.05,
    "mobile-phone": 0.05,
}


def linear_slope(values: List[float]) -> float:
    n = len(values)
    if n < 2:
        return 0.0
    xs = list(range(n))
    x_mean = sum(xs) / n
    y_mean = sum(values) / n
    num = sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, values))
    den = sum((x - x_mean) ** 2 for x in xs)
    return round(num / den, 2) if den else 0.0


def import_networth_from_excel(super_aud: float) -> int:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb["Historical"]

    existing = {row["date"]: row for row in read_csv(NETWORTH_CSV)}

    imported = 0
    for row in ws.iter_rows():
        vals = [cell.value for cell in row]
        if not vals or not isinstance(vals[0], dt_class):
            continue
        date_str = vals[0].strftime("%Y-%m-%d")
        everyday = float(vals[1] or 0)
        savings = float(vals[2] or 0)
        selfwealth = float(vals[3] or 0)
        ibkr = float(vals[4] or 0)
        cash = round(everyday + savings, 2)
        investments = round(selfwealth + ibkr, 2)
        total = round(cash + investments, 2)
        existing[date_str] = {
            "date": date_str,
            "cash_aud": str(cash),
            "investments_aud": str(investments),
            "super_aud": "0",
            "total_aud": str(total),
        }
        imported += 1

    wb.close()

    if existing:
        latest_date = max(existing.keys())
        row = existing[latest_date]
        row["super_aud"] = str(round(super_aud, 2))
        row["total_aud"] = str(round(
            float(row["cash_aud"]) + float(row["investments_aud"]) + super_aud, 2
        ))

    sorted_rows = sorted(existing.values(), key=lambda r: r["date"])
    write_csv(NETWORTH_CSV, NETWORTH_FIELDS, sorted_rows)
    return imported


@app.get("/")
def root():
    return redirect("/dashboard")


@app.get("/dashboard")
def dashboard_page():
    return send_file(BASE_DIR / "dashboard.html")


@app.get("/bills")
def bills_page():
    return send_file(BASE_DIR / "bills.html")


@app.get("/spending")
def spending_page():
    return send_file(BASE_DIR / "spending.html")


@app.get("/insights")
def insights_page():
    return send_file(BASE_DIR / "insights.html")


@app.get("/api/status")
def api_status():
    return jsonify(get_status())


@app.get("/sync")
def sync_route():
    full_refresh = request.args.get("full", "").lower() in {"1", "true", "yes"}
    counts = sync_transactions(full_refresh=full_refresh)
    return jsonify({"ok": True, "counts": counts})


@app.get("/api/bills")
def api_bills():
    return jsonify(compute_bill_statuses())


@app.get("/api/bills/history")
def api_bills_history():
    return jsonify(compute_bill_history())


@app.get("/api/bills/cycle/<cycle_tag>")
def api_bill_cycle_detail(cycle_tag: str):
    history = compute_bill_history()
    cycle = next((h for h in history if h["cycle_tag"] == cycle_tag), None)
    if not cycle:
        return jsonify({"error": "Cycle not found"}), 404

    parsed = parse_cycle_tag(cycle_tag)
    if not parsed:
        return jsonify({"error": "Invalid cycle tag"}), 400
    slug, month, year = parsed

    bill_model = cycle_bill_model(slug, month, str(year))
    housemates = read_housemates()
    overrides = build_override_lookup()

    spending_rows = read_csv(DATA_DIR / "transactions_spending.csv")
    beem_rows = [r for r in spending_rows if is_incoming_beem(r)]

    total_due = cycle.get("total_due")

    housemate_detail = []
    for hm in housemates:
        name = (hm.get("name") or "").lower()
        share = share_for_housemate(bill_model, hm, total_due)
        paid = False
        source = None
        paid_date = None

        for txn in beem_rows:
            tags = tag_set(txn.get("tags", ""))
            expanded = expand_bill_cycle_tags(tags)
            if name in tags and cycle_tag in expanded:
                paid = True
                source = "tag"
                paid_date = txn.get("settled_at") or txn.get("created_at")
                break

        if not paid:
            override = overrides.get((name, slug.lower(), month, str(year)))
            if override is not None:
                paid = parse_bool(override.get("paid"))
                source = "manual" if paid else None
                paid_date = override.get("paid_date") or None

        housemate_detail.append({
            "name": name,
            "share": share,
            "paid": paid,
            "source": source,
            "paid_date": paid_date,
        })

    timeline = []
    if cycle.get("paid_date"):
        timeline.append({
            "date": cycle["paid_date"],
            "event": f"Bill paid to {cycle.get('provider') or slug}",
            "type": "payment_out",
        })
    for hm in housemate_detail:
        if hm["paid"] and hm["paid_date"]:
            share_str = f"${hm['share']:.2f}" if hm["share"] else "share"
            timeline.append({
                "date": hm["paid_date"],
                "event": f"{hm['name'].title()} paid {share_str}",
                "type": "housemate_paid",
            })
    timeline.sort(key=lambda e: e["date"] or "")

    config = load_config()
    notes = config.get(f"bill_notes_{cycle_tag}", "")

    return jsonify({
        "cycle_tag": cycle_tag,
        "slug": slug,
        "month": month,
        "year": int(year),
        "label": cycle["label"],
        "provider": cycle.get("provider", ""),
        "status": cycle["status"],
        "paid_date": cycle.get("paid_date"),
        "total_due": total_due,
        "incoming_total": cycle.get("incoming_total", 0),
        "housemates": housemate_detail,
        "timeline": timeline,
        "notes": notes,
    })


@app.post("/api/bills/cycle/<cycle_tag>/override")
def api_bill_cycle_override(cycle_tag: str):
    parsed = parse_cycle_tag(cycle_tag)
    if not parsed:
        return jsonify({"ok": False, "error": "Invalid cycle tag"}), 400
    slug, month, year = parsed

    payload = request.get_json(force=True) or {}
    housemate = (payload.get("housemate") or "").lower().strip()
    if housemate not in HOUSEMATE_NAMES:
        return jsonify({"ok": False, "error": "Unknown housemate"}), 400

    paid = bool(payload.get("paid"))
    note = payload.get("note", "")
    paid_date = payload.get("paid_date", "")

    overrides = read_manual_overrides()
    fieldnames = ["housemate", "slug", "month", "year", "paid", "note", "paid_date"]

    updated = False
    for row in overrides:
        if (
            row.get("housemate", "").lower() == housemate
            and row.get("slug", "").lower() == slug.lower()
            and row.get("month", "").lower() == month
            and str(row.get("year", "")) == str(year)
        ):
            row["paid"] = "true" if paid else "false"
            row["note"] = note
            row["paid_date"] = paid_date
            updated = True
            break

    if not updated:
        overrides.append({
            "housemate": housemate,
            "slug": slug,
            "month": month,
            "year": str(year),
            "paid": "true" if paid else "false",
            "note": note,
            "paid_date": paid_date,
        })

    write_csv(DATA_DIR / "manual_overrides.csv", fieldnames, overrides)
    return jsonify({"ok": True, "housemate": housemate, "paid": paid})


@app.post("/api/bills/cycle/<cycle_tag>/notes")
def api_bill_cycle_notes(cycle_tag: str):
    parsed = parse_cycle_tag(cycle_tag)
    if not parsed:
        return jsonify({"ok": False, "error": "Invalid cycle tag"}), 400

    payload = request.get_json(force=True) or {}
    notes = (payload.get("notes") or "").strip()

    config = load_config()
    config[f"bill_notes_{cycle_tag}"] = notes
    CONFIG_PATH.write_text(json.dumps(config, indent=2))
    return jsonify({"ok": True, "notes": notes})


@app.post("/api/bills")
def api_add_bill():
    payload = request.get_json(force=True) or {}
    required = ["slug", "label", "due_date", "recurrence", "split_type"]
    missing = [field for field in required if not payload.get(field)]
    if missing:
        return jsonify({"ok": False, "error": f"Missing fields: {', '.join(missing)}"}), 400
    if payload["slug"] not in BILL_SLUGS:
        return jsonify({"ok": False, "error": "Invalid bill slug"}), 400
    row = append_bill(payload)
    return jsonify(row), 201


@app.post("/api/bills/<int:bill_id>/override")
def api_override_bill(bill_id: int):
    payload = request.get_json(force=True) or {}
    try:
        row = upsert_override(bill_id, payload)
    except KeyError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    return jsonify({"ok": True, "override": row})


@app.get("/api/spending")
def api_spending():
    since = request.args.get("since")
    until = request.args.get("until")
    category = request.args.get("category")
    rows = read_csv(DATA_DIR / "transactions_spending.csv")
    return jsonify(filter_spending_rows(rows, since, until, category))


def _cashflow_monthly():
    months_back = 12
    today = datetime.now()
    month_ranges = []
    d = today.replace(day=1)
    for _ in range(months_back):
        d = (d - timedelta(days=1)).replace(day=1)
        next_d = (d.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_ranges.append((d, next_d, f"{d.year}-{d.month:02d}"))
    month_ranges.reverse()

    status = get_status()
    account_ids = status.get("account_ids", {})
    savings_id = account_ids.get("savings", "")
    grow_id = account_ids.get("grow", "")
    two_up_id = account_ids.get("two_up", "")
    spending_id = account_ids.get("spending", "")
    internal_ids_savings = {tid for tid in (spending_id, grow_id, two_up_id) if tid}

    INVESTMENT_DESCRIPTIONS = {"ibkr", "selfwealth", "stake", "commsec", "pearler"}

    spending_rows = read_csv(DATA_DIR / "transactions_spending.csv")
    savings_csv = DATA_DIR / "transactions_savings.csv"
    savings_rows = read_csv(savings_csv) if savings_csv.exists() else []

    result = []
    for start, end, label in month_ranges:
        income = 0.0
        saved = 0.0
        invested = 0.0
        disc = 0.0

        for row in spending_rows:
            dt_str = row.get("settled_at") or row.get("created_at") or ""
            if not dt_str:
                continue
            dt = parse_datetime_or_date(dt_str)
            if not (start <= dt < end):
                continue
            amount = parse_float(row.get("amount")) or 0.0
            tid = row.get("transfer_account_id", "")
            if amount > 0 and not tid:
                income += amount
            elif tid == savings_id and amount < 0:
                saved += abs(amount)
            elif tid == grow_id and amount < 0:
                saved += abs(amount)
            elif amount < 0 and not tid:
                disc += abs(amount)

        for row in savings_rows:
            dt_str = row.get("settled_at") or row.get("created_at") or ""
            if not dt_str:
                continue
            dt = parse_datetime_or_date(dt_str)
            if not (start <= dt < end):
                continue
            amount = parse_float(row.get("amount")) or 0.0
            if amount >= 0:
                continue
            tid = row.get("transfer_account_id", "")
            if tid in internal_ids_savings:
                continue
            desc = (row.get("description") or "").lower()
            if any(kw in desc for kw in INVESTMENT_DESCRIPTIONS):
                invested += abs(amount)

        save_rate = round((saved + invested) / income * 100, 1) if income > 0 else 0.0
        result.append({
            "month": label,
            "income": round(income, 2),
            "saved": round(saved, 2),
            "invested": round(invested, 2),
            "discretionary": round(disc, 2),
            "savings_rate": save_rate,
        })

    return jsonify(result)


@app.get("/api/spending/cashflow")
def api_spending_cashflow():
    if request.args.get("monthly", "").lower() in {"1", "true", "yes"}:
        return _cashflow_monthly()
    since = request.args.get("since")
    until = request.args.get("until")

    status = get_status()
    account_ids = status.get("account_ids", {})
    savings_id = account_ids.get("savings", "")
    grow_id = account_ids.get("grow", "")
    two_up_id = account_ids.get("two_up", "")

    since_dt = parser.isoparse(since) if since else None
    until_dt = parser.isoparse(until) if until else None
    if until_dt and until_dt.hour == 0 and until_dt.minute == 0 and until_dt.second == 0:
        until_dt = until_dt + timedelta(days=1)

    INVESTMENT_DESCRIPTIONS = {"ibkr", "selfwealth", "stake", "commsec", "pearler"}
    INSURANCE_DESCRIPTIONS = {"john symons"}
    TAX_DESCRIPTIONS = {"tax office payments", "australian taxation office", "ato"}

    income = 0.0
    savings_transfers = 0.0
    grow_transfers = 0.0
    two_up_transfers = 0.0
    investment_transfers = 0.0
    insurance_payments = 0.0
    tax_payments = 0.0
    discretionary = 0.0

    def _in_range(row: Dict) -> bool:
        created_at = row.get("settled_at") or row.get("created_at")
        if not created_at:
            return False
        dt = parse_datetime_or_date(created_at)
        if since_dt and dt < since_dt.replace(tzinfo=None):
            return False
        if until_dt and dt >= until_dt.replace(tzinfo=None):
            return False
        return True

    # Spending account: income in, internal transfers out, discretionary spend
    for row in read_csv(DATA_DIR / "transactions_spending.csv"):
        if not _in_range(row):
            continue
        amount = parse_float(row.get("amount")) or 0.0
        transfer_account_id = row.get("transfer_account_id", "")
        description = (row.get("description") or "").strip().lower()
        is_investment = any(kw in description for kw in INVESTMENT_DESCRIPTIONS)

        if amount > 0 and not transfer_account_id:
            income += amount
        elif transfer_account_id == savings_id and amount < 0:
            savings_transfers += abs(amount)
        elif transfer_account_id == grow_id and amount < 0:
            grow_transfers += abs(amount)
        elif transfer_account_id == two_up_id and amount < 0:
            two_up_transfers += abs(amount)
        elif is_investment and amount < 0:
            investment_transfers += abs(amount)
        elif amount < 0 and not transfer_account_id:
            discretionary += abs(amount)

    # Savings account: external outgoing payments (e.g. IBKR) — ignore internal Up transfers
    savings_csv = DATA_DIR / "transactions_savings.csv"
    if savings_csv.exists():
        spending_id = account_ids.get("spending", "")
        internal_ids = {tid for tid in (spending_id, grow_id, two_up_id) if tid}
        for row in read_csv(savings_csv):
            if not _in_range(row):
                continue
            amount = parse_float(row.get("amount")) or 0.0
            if amount >= 0:
                continue
            transfer_account_id = row.get("transfer_account_id", "")
            if transfer_account_id in internal_ids:
                continue  # internal Up transfer, already counted above
            description = (row.get("description") or "").strip().lower()
            if any(kw in description for kw in INVESTMENT_DESCRIPTIONS):
                investment_transfers += abs(amount)
            elif any(kw in description for kw in INSURANCE_DESCRIPTIONS):
                insurance_payments += abs(amount)
            elif any(kw in description for kw in TAX_DESCRIPTIONS):
                tax_payments += abs(amount)
            # other one-offs (Russell Barker, tipping pools etc.) are ignored

    total_out = savings_transfers + grow_transfers + investment_transfers + insurance_payments + tax_payments + discretionary
    return jsonify({
        "income": round(income, 2),
        "savings_transfers": round(savings_transfers, 2),
        "grow_transfers": round(grow_transfers, 2),
        "two_up_transfers": round(two_up_transfers, 2),
        "investment_transfers": round(investment_transfers, 2),
        "insurance_payments": round(insurance_payments, 2),
        "tax_payments": round(tax_payments, 2),
        "discretionary": round(discretionary, 2),
        "net": round(income - two_up_transfers - total_out, 2),
    })


@app.get("/api/budgets")
def api_get_budgets():
    return jsonify(get_budgets())


@app.post("/api/budgets")
def api_save_budgets():
    payload = request.get_json(force=True) or {}
    budgets = {}
    for k, v in payload.items():
        try:
            val = float(v)
            if val >= 0:
                budgets[str(k)] = val
        except (TypeError, ValueError):
            pass
    save_budgets(budgets)
    return jsonify({"ok": True, "budgets": budgets})


def get_recurring_exclusions() -> set:
    if not CONFIG_PATH.exists():
        return set()
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    return set(config.get("recurring_exclusions", []))


def save_recurring_exclusions(exclusions: set) -> None:
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    config["recurring_exclusions"] = sorted(exclusions)
    with CONFIG_PATH.open("w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2)
        handle.write("\n")


@app.get("/api/spending/recurring")
def api_spending_recurring():
    rows = read_csv(DATA_DIR / "transactions_spending.csv")
    exclusions = get_recurring_exclusions()
    return jsonify(detect_recurring(rows, exclusions))


@app.post("/api/spending/recurring/exclude")
def api_recurring_exclude():
    payload = request.get_json(force=True) or {}
    description = (payload.get("description") or "").strip()
    if not description:
        return jsonify({"ok": False, "error": "description required"}), 400
    exclusions = get_recurring_exclusions()
    exclusions.add(description)
    save_recurring_exclusions(exclusions)
    return jsonify({"ok": True, "excluded": description})


@app.get("/api/spending/category-history")
def api_spending_category_history():
    months_back = int(request.args.get("months", 6))
    rows = read_csv(DATA_DIR / "transactions_spending.csv")

    status = get_status()
    account_ids = status.get("account_ids", {})
    two_up_id = account_ids.get("two_up", "")
    savings_id = account_ids.get("savings", "")
    grow_id = account_ids.get("grow", "")
    internal_ids = {tid for tid in (two_up_id, savings_id, grow_id) if tid}

    today = datetime.now()
    cutoff = today.replace(day=1)
    for _ in range(months_back):
        cutoff = (cutoff - timedelta(days=1)).replace(day=1)

    months: List[str] = []
    d = cutoff
    current_month = today.replace(day=1)
    while d < current_month:
        months.append(f"{d.year}-{d.month:02d}")
        next_month = d.month + 1 if d.month < 12 else 1
        next_year = d.year + (1 if d.month == 12 else 0)
        d = d.replace(year=next_year, month=next_month)

    spend: Dict[str, Dict[str, float]] = {}
    for row in rows:
        amount = parse_float(row.get("amount")) or 0.0
        if amount >= 0:
            continue
        transfer_id = row.get("transfer_account_id", "")
        if transfer_id in internal_ids:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        dt = parse_datetime_or_date(dt_str)
        month_key = f"{dt.year}-{dt.month:02d}"
        if month_key not in months:
            continue
        cat = (row.get("category") or "").strip() or "uncategorised"
        spend.setdefault(month_key, {})
        spend[month_key][cat] = spend[month_key].get(cat, 0.0) + abs(amount)

    cat_totals: Dict[str, float] = {}
    for month_data in spend.values():
        for cat, total in month_data.items():
            cat_totals[cat] = cat_totals.get(cat, 0.0) + total
    top_cats = sorted(cat_totals, key=lambda c: cat_totals[c], reverse=True)[:12]

    result = []
    for cat in top_cats:
        monthly_values = [
            {"month": month, "total": round(spend.get(month, {}).get(cat, 0.0), 2)}
            for month in months
        ]
        result.append({
            "category": cat,
            "total": round(cat_totals[cat], 2),
            "monthly": monthly_values,
        })

    return jsonify({"months": months, "categories": result})


def _get_month_range(month_str: str) -> tuple:
    """Return (start_dt, end_dt) for a YYYY-MM string."""
    year, month = int(month_str[:4]), int(month_str[5:7])
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    return start, end


def _compute_month_cashflow(month_str: str) -> Dict:
    """Return income, saved, discretionary for a single calendar month."""
    start, end = _get_month_range(month_str)
    status = get_status()
    account_ids = status.get("account_ids", {})
    savings_id = account_ids.get("savings", "")
    grow_id = account_ids.get("grow", "")
    two_up_id = account_ids.get("two_up", "")
    spending_id = account_ids.get("spending", "")
    internal_ids_savings = {tid for tid in (spending_id, grow_id, two_up_id) if tid}

    INVESTMENT_DESCRIPTIONS = {"ibkr", "selfwealth", "stake", "commsec", "pearler"}

    # Build a multiset of (date, amount) for 2Up transfers out, used to identify
    # Beem credits that are just pass-throughs forwarded straight to 2Up.
    all_rows = read_csv(DATA_DIR / "transactions_spending.csv")
    two_up_transfers: Dict[str, List[float]] = {}
    for row in all_rows:
        if row.get("transfer_account_id", "") != two_up_id:
            continue
        amt = parse_float(row.get("amount")) or 0.0
        if amt >= 0:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        try:
            dt = parse_datetime_or_date(dt_str)
        except Exception:
            continue
        day = dt.strftime("%Y-%m-%d")
        two_up_transfers.setdefault(day, []).append(round(abs(amt), 2))

    def _is_passthrough_beem(beem_date: datetime, beem_amount: float) -> bool:
        """True if a matching 2Up transfer exists within 2 days of this Beem credit."""
        amt_r = round(beem_amount, 2)
        for delta in range(3):
            day = (beem_date + timedelta(days=delta)).strftime("%Y-%m-%d")
            pool = two_up_transfers.get(day, [])
            if amt_r in pool:
                pool.remove(amt_r)
                return True
        return False

    income = 0.0
    saved = 0.0
    invested = 0.0
    disc = 0.0
    reimbursements = 0.0

    for row in all_rows:
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        dt = parse_datetime_or_date(dt_str)
        if not (start <= dt < end):
            continue
        amount = parse_float(row.get("amount")) or 0.0
        tid = row.get("transfer_account_id", "")
        desc = (row.get("description") or "").strip().lower()
        is_beem = desc == "beem"
        if amount > 0 and not tid:
            if is_beem and not _is_passthrough_beem(dt, amount):
                reimbursements += amount
            elif not is_beem:
                income += amount
        elif tid == savings_id and amount < 0:
            saved += abs(amount)
        elif tid == grow_id and amount < 0:
            saved += abs(amount)
        elif amount < 0 and not tid:
            disc += abs(amount)

    savings_csv = DATA_DIR / "transactions_savings.csv"
    if savings_csv.exists():
        for row in read_csv(savings_csv):
            dt_str = row.get("settled_at") or row.get("created_at") or ""
            if not dt_str:
                continue
            dt = parse_datetime_or_date(dt_str)
            if not (start <= dt < end):
                continue
            amount = parse_float(row.get("amount")) or 0.0
            if amount >= 0:
                continue
            tid = row.get("transfer_account_id", "")
            if tid in internal_ids_savings:
                continue
            desc = (row.get("description") or "").lower()
            if any(kw in desc for kw in INVESTMENT_DESCRIPTIONS):
                invested += abs(amount)

    net_disc = max(disc - reimbursements, 0.0)
    savings_rate = round((saved + invested) / income * 100, 1) if income > 0 else 0.0
    return {
        "income": round(income, 2),
        "saved": round(saved + invested, 2),
        "discretionary": round(net_disc, 2),
        "reimbursements": round(reimbursements, 2),
        "savings_rate": savings_rate,
    }


@app.get("/api/insights/monthly")
def api_insights_monthly():
    today = datetime.now()
    first_of_this_month = today.replace(day=1)
    default_month_dt = (first_of_this_month - timedelta(days=1)).replace(day=1)
    default_month = f"{default_month_dt.year}-{default_month_dt.month:02d}"
    month_str = request.args.get("month", default_month)

    try:
        year, month = int(month_str[:4]), int(month_str[5:7])
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            raise ValueError
    except (ValueError, IndexError):
        return jsonify({"error": "invalid month format, use YYYY-MM"}), 400

    prev_dt = datetime(year, month, 1) - timedelta(days=1)
    prev_month = f"{prev_dt.year}-{prev_dt.month:02d}"

    all_rows = read_csv(DATA_DIR / "transactions_spending.csv")
    month_set: set = set()
    for row in all_rows:
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        try:
            dt = parse_datetime_or_date(dt_str)
            mk = f"{dt.year}-{dt.month:02d}"
            if mk <= default_month:
                month_set.add(mk)
        except Exception:
            continue
    available_months = sorted(month_set, reverse=True)

    cf = _compute_month_cashflow(month_str)
    cf_prev = _compute_month_cashflow(prev_month)

    start, end = _get_month_range(month_str)
    prev_start, prev_end = _get_month_range(prev_month)

    status = get_status()
    account_ids = status.get("account_ids", {})
    two_up_id = account_ids.get("two_up", "")
    savings_id = account_ids.get("savings", "")
    grow_id = account_ids.get("grow", "")
    internal_ids = {tid for tid in (two_up_id, savings_id, grow_id) if tid}

    cat_spend: Dict[str, float] = {}
    cat_spend_prev: Dict[str, float] = {}
    merchant_totals: Dict[str, float] = {}
    merchant_counts: Dict[str, int] = {}
    merchant_cats: Dict[str, str] = {}

    for row in all_rows:
        amount = parse_float(row.get("amount")) or 0.0
        if amount >= 0:
            continue
        tid = row.get("transfer_account_id", "")
        if tid in internal_ids:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        try:
            dt = parse_datetime_or_date(dt_str)
        except Exception:
            continue
        cat = (row.get("category") or "").strip() or "uncategorised"
        desc = (row.get("description") or "").strip()

        if start <= dt < end:
            cat_spend[cat] = cat_spend.get(cat, 0.0) + abs(amount)
            if desc:
                merchant_totals[desc] = merchant_totals.get(desc, 0.0) + abs(amount)
                merchant_counts[desc] = merchant_counts.get(desc, 0) + 1
                if desc not in merchant_cats:
                    merchant_cats[desc] = cat
        elif prev_start <= dt < prev_end:
            cat_spend_prev[cat] = cat_spend_prev.get(cat, 0.0) + abs(amount)

    disc = cf["discretionary"] or 1.0

    top_categories = sorted(
        [
            {
                "slug": cat,
                "total": round(total, 2),
                "prev_total": round(cat_spend_prev.get(cat, 0.0), 2),
                "pct_of_disc": round(total / disc * 100, 1),
            }
            for cat, total in cat_spend.items()
        ],
        key=lambda x: x["total"],
        reverse=True,
    )[:15]

    sorted_merchants = sorted(merchant_totals.items(), key=lambda x: x[1], reverse=True)[:15]
    cumulative = 0.0
    top_merchants = []
    for desc, total in sorted_merchants:
        cumulative += total
        top_merchants.append({
            "description": desc,
            "total": round(total, 2),
            "count": merchant_counts[desc],
            "cumulative": round(cumulative, 2),
            "category": merchant_cats.get(desc, ""),
        })

    weighted_total = sum(
        cat_spend.get(cat, 0.0) * FRIVOLITY_WEIGHTS.get(cat, 0.5)
        for cat in cat_spend
    )
    frivolity_score = round(weighted_total / disc * 100, 1)
    drivers = sorted(
        [
            {
                "slug": cat,
                "total": round(cat_spend[cat], 2),
                "weight": FRIVOLITY_WEIGHTS.get(cat, 0.5),
                "contribution": round(cat_spend[cat] * FRIVOLITY_WEIGHTS.get(cat, 0.5), 2),
            }
            for cat in cat_spend
            if cat_spend[cat] * FRIVOLITY_WEIGHTS.get(cat, 0.5) > 0
        ],
        key=lambda x: x["contribution"],
        reverse=True,
    )[:8]

    frivolity_history = []
    hist_months: List[str] = []
    d = default_month_dt
    for _ in range(6):
        hist_months.append(f"{d.year}-{d.month:02d}")
        d = (d - timedelta(days=1)).replace(day=1)
    hist_months.reverse()

    for hm in hist_months:
        hs, he = _get_month_range(hm)
        hm_cat: Dict[str, float] = {}
        hm_disc = 0.0
        for row in all_rows:
            amount = parse_float(row.get("amount")) or 0.0
            if amount >= 0:
                continue
            tid = row.get("transfer_account_id", "")
            if tid in internal_ids:
                continue
            dt_str = row.get("settled_at") or row.get("created_at") or ""
            if not dt_str:
                continue
            try:
                dt = parse_datetime_or_date(dt_str)
            except Exception:
                continue
            if hs <= dt < he:
                cat = (row.get("category") or "").strip() or "uncategorised"
                hm_cat[cat] = hm_cat.get(cat, 0.0) + abs(amount)
                hm_disc += abs(amount)
        hm_weighted = sum(hm_cat.get(c, 0.0) * FRIVOLITY_WEIGHTS.get(c, 0.5) for c in hm_cat)
        hm_score = round(hm_weighted / hm_disc * 100, 1) if hm_disc > 0 else 0.0
        frivolity_history.append({"month": hm, "score": hm_score})

    cat_monthly: Dict[str, List[float]] = {}
    for hm in hist_months:
        hs, he = _get_month_range(hm)
        hm_cat2: Dict[str, float] = {}
        for row in all_rows:
            amount = parse_float(row.get("amount")) or 0.0
            if amount >= 0:
                continue
            tid = row.get("transfer_account_id", "")
            if tid in internal_ids:
                continue
            dt_str = row.get("settled_at") or row.get("created_at") or ""
            if not dt_str:
                continue
            try:
                dt = parse_datetime_or_date(dt_str)
            except Exception:
                continue
            if hs <= dt < he:
                cat = (row.get("category") or "").strip() or "uncategorised"
                hm_cat2[cat] = hm_cat2.get(cat, 0.0) + abs(amount)
        for cat in set(list(cat_monthly.keys()) + list(hm_cat2.keys())):
            if cat not in cat_monthly:
                cat_monthly[cat] = [0.0] * len(hist_months)
            cat_monthly[cat][hist_months.index(hm)] = hm_cat2.get(cat, 0.0)

    all_cat_totals = {cat: sum(vals) for cat, vals in cat_monthly.items()}
    top_trend_cats = sorted(all_cat_totals, key=lambda c: all_cat_totals[c], reverse=True)[:8]
    trends = []
    for cat in top_trend_cats:
        vals = cat_monthly[cat]
        recent_3 = [v for v in vals[-3:] if v > 0]
        avg_3mo = round(sum(recent_3) / len(recent_3), 2) if recent_3 else 0.0
        this_month_val = cat_spend.get(cat, 0.0)
        trends.append({
            "slug": cat,
            "slope_per_month": linear_slope(vals),
            "avg_3mo": avg_3mo,
            "this_month": round(this_month_val, 2),
        })

    return jsonify({
        "month": month_str,
        "prev_month": prev_month,
        "available_months": available_months,
        "income": cf["income"],
        "saved": cf["saved"],
        "discretionary": cf["discretionary"],
        "reimbursements": cf["reimbursements"],
        "savings_rate": cf["savings_rate"],
        "prev_savings_rate": cf_prev["savings_rate"],
        "prev_discretionary": cf_prev["discretionary"],
        "top_categories": top_categories,
        "top_merchants": top_merchants,
        "frivolity": {
            "score": frivolity_score,
            "weighted_total": round(weighted_total, 2),
            "drivers": drivers,
            "history": frivolity_history,
        },
        "trends": trends,
    })


def get_subscription_tags() -> Dict[str, str]:
    if not CONFIG_PATH.exists():
        return {}
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    return config.get("subscription_tags", {})


def save_subscription_tag(description: str, tag: Optional[str]) -> None:
    with CONFIG_PATH.open("r", encoding="utf-8") as handle:
        config = json.load(handle)
    tags = config.get("subscription_tags", {})
    if tag is None:
        tags.pop(description, None)
    else:
        tags[description] = tag
    config["subscription_tags"] = tags
    with CONFIG_PATH.open("w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2)
        handle.write("\n")


@app.get("/api/insights/subscription-tags")
def api_get_subscription_tags():
    return jsonify(get_subscription_tags())


@app.post("/api/insights/subscription-tag")
def api_subscription_tag():
    payload = request.get_json(force=True) or {}
    description = (payload.get("description") or "").strip()
    tag = payload.get("tag")
    if not description:
        return jsonify({"ok": False, "error": "description required"}), 400
    if tag is not None and tag not in {"KEEP", "REVIEW", "CUT"}:
        return jsonify({"ok": False, "error": "tag must be KEEP, REVIEW, CUT, or null"}), 400
    save_subscription_tag(description, tag)
    return jsonify({"ok": True, "description": description, "tag": tag})


@app.get("/api/insights/category")
def api_insights_category():
    slug = (request.args.get("slug") or "").strip()
    if not slug:
        return jsonify({"error": "slug required"}), 400
    month_str = request.args.get("month", "")
    if not month_str:
        today = datetime.now()
        first_of_this_month = today.replace(day=1)
        default_dt = (first_of_this_month - timedelta(days=1)).replace(day=1)
        month_str = f"{default_dt.year}-{default_dt.month:02d}"
    try:
        year, month = int(month_str[:4]), int(month_str[5:7])
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            raise ValueError
    except (ValueError, IndexError):
        return jsonify({"error": "invalid month"}), 400

    status = get_status()
    account_ids = status.get("account_ids", {})
    internal_ids = {
        tid for tid in (
            account_ids.get("two_up", ""),
            account_ids.get("savings", ""),
            account_ids.get("grow", ""),
        ) if tid
    }

    # Build 6-month history window ending at month_str
    hist_months: List[str] = []
    d = datetime(year, month, 1)
    for _ in range(6):
        hist_months.append(f"{d.year}-{d.month:02d}")
        d = (d - timedelta(days=1)).replace(day=1)
    hist_months.reverse()

    all_rows = read_csv(DATA_DIR / "transactions_spending.csv")

    # Selected month range
    sel_start, sel_end = _get_month_range(month_str)

    history: Dict[str, float] = {m: 0.0 for m in hist_months}
    merchant_totals: Dict[str, float] = {}
    merchant_counts: Dict[str, int] = {}
    transactions: List[Dict] = []

    for row in all_rows:
        amount = parse_float(row.get("amount")) or 0.0
        if amount >= 0:
            continue
        if row.get("transfer_account_id", "") in internal_ids:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        try:
            dt = parse_datetime_or_date(dt_str)
        except Exception:
            continue
        cat = (row.get("category") or "").strip() or "uncategorised"
        if cat != slug:
            continue
        mk = f"{dt.year}-{dt.month:02d}"
        if mk in history:
            history[mk] = history[mk] + abs(amount)
        if sel_start <= dt < sel_end:
            desc = (row.get("description") or "").strip()
            merchant_totals[desc] = merchant_totals.get(desc, 0.0) + abs(amount)
            merchant_counts[desc] = merchant_counts.get(desc, 0) + 1
            transactions.append({
                "date": dt.strftime("%Y-%m-%d"),
                "description": desc,
                "category": cat,
                "amount": round(abs(amount), 2),
            })

    transactions.sort(key=lambda x: x["date"], reverse=True)
    top_merchants = sorted(
        [
            {"description": d, "total": round(t, 2), "count": merchant_counts[d]}
            for d, t in merchant_totals.items()
        ],
        key=lambda x: x["total"],
        reverse=True,
    )[:10]

    return jsonify({
        "slug": slug,
        "month": month_str,
        "history": [{"month": m, "total": round(history[m], 2)} for m in hist_months],
        "top_merchants": top_merchants,
        "transactions": transactions,
    })


@app.get("/api/insights/merchant")
def api_insights_merchant():
    name = (request.args.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    month_str = request.args.get("month", "")
    if not month_str:
        today = datetime.now()
        first_of_this_month = today.replace(day=1)
        default_dt = (first_of_this_month - timedelta(days=1)).replace(day=1)
        month_str = f"{default_dt.year}-{default_dt.month:02d}"
    try:
        year, month = int(month_str[:4]), int(month_str[5:7])
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            raise ValueError
    except (ValueError, IndexError):
        return jsonify({"error": "invalid month"}), 400

    status = get_status()
    account_ids = status.get("account_ids", {})
    internal_ids = {
        tid for tid in (
            account_ids.get("two_up", ""),
            account_ids.get("savings", ""),
            account_ids.get("grow", ""),
        ) if tid
    }

    hist_months: List[str] = []
    d = datetime(year, month, 1)
    for _ in range(6):
        hist_months.append(f"{d.year}-{d.month:02d}")
        d = (d - timedelta(days=1)).replace(day=1)
    hist_months.reverse()

    all_rows = read_csv(DATA_DIR / "transactions_spending.csv")
    sel_start, sel_end = _get_month_range(month_str)

    history: Dict[str, float] = {m: 0.0 for m in hist_months}
    transactions: List[Dict] = []

    for row in all_rows:
        amount = parse_float(row.get("amount")) or 0.0
        if amount >= 0:
            continue
        if row.get("transfer_account_id", "") in internal_ids:
            continue
        dt_str = row.get("settled_at") or row.get("created_at") or ""
        if not dt_str:
            continue
        try:
            dt = parse_datetime_or_date(dt_str)
        except Exception:
            continue
        desc = (row.get("description") or "").strip()
        if desc != name:
            continue
        mk = f"{dt.year}-{dt.month:02d}"
        if mk in history:
            history[mk] = history[mk] + abs(amount)
        if sel_start <= dt < sel_end:
            cat = (row.get("category") or "").strip() or "uncategorised"
            transactions.append({
                "date": dt.strftime("%Y-%m-%d"),
                "description": desc,
                "category": cat,
                "amount": round(abs(amount), 2),
            })

    transactions.sort(key=lambda x: x["date"], reverse=True)

    return jsonify({
        "name": name,
        "month": month_str,
        "history": [{"month": m, "total": round(history[m], 2)} for m in hist_months],
        "transactions": transactions,
    })


@app.get("/api/spending/summary")
def api_spending_summary():
    since = request.args.get("since")
    until = request.args.get("until")
    category = request.args.get("category")
    group_by = request.args.get("group_by", "month")
    exclude_refunds = request.args.get("exclude_refunds", "").lower() in {"1", "true", "yes"}
    min_amount = request.args.get("min_amount")

    rows = read_csv(DATA_DIR / "transactions_spending.csv")
    filtered = filter_spending_rows(rows, since, until, category)

    if exclude_refunds:
        filtered = [row for row in filtered if float(row.get("amount", 0)) < 0]

    if min_amount:
        try:
            threshold = abs(float(min_amount))
            filtered = [row for row in filtered if abs(float(row.get("amount", 0))) >= threshold]
        except ValueError:
            pass

    if group_by not in {"week", "month"}:
        group_by = "month"

    grouped = group_spending_by_period(filtered, group_by)

    merchant_totals: Dict[str, float] = {}
    merchant_counts: Dict[str, int] = {}
    for row in filtered:
        amount = float(row.get("amount", 0))
        if amount >= 0:
            continue
        desc = (row.get("description") or "").strip()
        if not desc:
            continue
        merchant_totals[desc] = merchant_totals.get(desc, 0.0) + abs(amount)
        merchant_counts[desc] = merchant_counts.get(desc, 0) + 1

    merchants = sorted(
        [
            {"description": desc, "total_spend": round(merchant_totals[desc], 2), "transaction_count": merchant_counts[desc]}
            for desc in merchant_totals
        ],
        key=lambda merchant: merchant["total_spend"],
        reverse=True,
    )[:20]

    total_spend = round(sum(abs(float(row.get("amount", 0))) for row in filtered if float(row.get("amount", 0)) < 0), 2)
    total_in = round(sum(float(row.get("amount", 0)) for row in filtered if float(row.get("amount", 0)) > 0), 2)

    return jsonify(
        {
            "group_by": group_by,
            "periods": grouped,
            "merchants": merchants,
            "total_spend": total_spend,
            "total_in": total_in,
            "transaction_count": len(filtered),
        }
    )


@app.get("/budget")
def budget_page():
    return send_file(BASE_DIR / "budget.html")


@app.get("/networth")
def networth_page():
    return send_file(BASE_DIR / "networth.html")


@app.get("/portfolio")
def portfolio_page():
    return send_file(BASE_DIR / "portfolio.html")


@app.get("/cgt")
def cgt_page():
    return send_file(BASE_DIR / "cgt.html")


@app.get("/house")
def house_page():
    return send_file(BASE_DIR / "house.html")


BILLS_UPLOAD_DIR = BASE_DIR / "imports" / "bills"
BILLS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


@app.post("/api/bills/upload")
def api_bills_upload():
    from flask import send_from_directory
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "Empty filename"}), 400
    safe_name = re.sub(r"[^\w.\- ]", "_", f.filename)
    dest = BILLS_UPLOAD_DIR / safe_name
    f.save(str(dest))
    return jsonify({"ok": True, "filename": safe_name, "path": str(dest)}), 201


@app.get("/api/bills/uploads")
def api_bills_uploads():
    files = []
    for p in sorted(BILLS_UPLOAD_DIR.iterdir()):
        if p.is_file():
            files.append({"filename": p.name, "size": p.stat().st_size, "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat()})
    return jsonify(files)


RECURRING_CSV = DATA_DIR / "recurring_investments.csv"
IBKR_INCEPTION = BASE_DIR / "imports" / "ibkr" / "Angus_J_Symons_Inception_May_15_2026.csv"
IBKR_FLEX = BASE_DIR / "imports" / "ibkr" / "U9504716_U9504716_20250516_20260515_AF_NA_47d9d492ad59520899300f36f9855bde.csv"


def parse_ibkr_performance():
    """Parse monthly TWR, dividends, and MTM from IBKR inception report + flex query."""
    import re as _re
    monthly_twr = {}
    dividends_by_year = {}
    mtm_by_ticker = {}

    if IBKR_INCEPTION.exists():
        with IBKR_INCEPTION.open(encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines:
            if line.startswith("Historical Performance Benchmark Comparison,Data,"):
                parts = line.strip().split(",")
                month = parts[2]
                if _re.match(r"^\d{6}$", month):
                    val = parts[-1].strip()
                    if val and val != "-":
                        try:
                            monthly_twr[month] = float(val)
                        except ValueError:
                            pass
            elif line.startswith("Dividends,Data"):
                parts = line.strip().split(",")
                try:
                    year = parts[2][:4]
                    amount = float(parts[7])
                    dividends_by_year[year] = dividends_by_year.get(year, 0.0) + amount
                except (ValueError, IndexError):
                    pass

    if IBKR_FLEX.exists():
        with IBKR_FLEX.open(encoding="utf-8") as f:
            content = f.read()
        lines = content.splitlines()
        in_mtm = False
        for line in lines:
            if "TransactionMtmPnl" in line and "Symbol" in line:
                in_mtm = True
                continue
            if in_mtm and not line.startswith('"'):
                in_mtm = False
            if in_mtm and line.startswith('"'):
                parts = [p.strip('"') for p in line.split('","')]
                if len(parts) >= 4 and parts[0] not in ("", "AUD", "USD"):
                    try:
                        mtm_by_ticker[parts[0]] = {
                            "transaction_mtm": float(parts[1]),
                            "prior_open_mtm": float(parts[2]),
                            "commissions": float(parts[3]),
                            "total": float(parts[4]),
                        }
                    except (ValueError, IndexError):
                        pass

    return {
        "monthly_twr": monthly_twr,
        "dividends_by_year": {k: round(v, 2) for k, v in dividends_by_year.items()},
        "mtm_by_ticker": mtm_by_ticker,
    }


@app.get("/api/performance")
def api_performance():
    return jsonify(parse_ibkr_performance())


@app.get("/api/recurring")
def api_recurring():
    return jsonify(read_csv(RECURRING_CSV))


@app.get("/performance")
def performance_page():
    return send_file(BASE_DIR / "performance.html")


@app.get("/tax")
def tax_page():
    return send_file(BASE_DIR / "tax.html")


HEALTH_DIR = BASE_DIR / "health"


@app.get("/health")
def health_root():
    return redirect("/health/anti-age")


@app.get("/health/<tab>")
def health_page(tab: str):
    page = HEALTH_DIR / f"{tab}.html"
    if not page.exists():
        return "Not found", 404
    return send_file(page)


# ── Health data API ──────────────────────────────────────────────────────────

from health_pipeline.metrics import (
    hrv_daily, resting_hr_daily, sleep_daily, spo2_daily, resp_rate_daily,
    steps_daily, vo2_trend, nutrition_daily, workout_sessions,
    import_status, get_config, set_config, hrv_baseline, resting_hr_baseline,
)
from health_pipeline.parse_macrofactor import (
    import_nutrition as _import_nutrition,
    import_workouts as _import_workouts,
    NUTRITION_PATH, WORKOUTS_PATH,
)
from health_pipeline.parse_health_json import parse_and_import, latest_health_json

HEALTH_UPLOAD_DIR = BASE_DIR / "imports" / "health"
HEALTH_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


@app.get("/api/health/status")
def api_health_status():
    return jsonify({
        "imports": import_status(),
        "config": get_config(),
    })


@app.get("/api/health/hrv")
def api_health_hrv():
    days = int(request.args.get("days", 30))
    return jsonify({
        "data": hrv_daily(days),
        "baseline": hrv_baseline(days),
    })


@app.get("/api/health/resting-hr")
def api_health_resting_hr():
    days = int(request.args.get("days", 30))
    return jsonify({
        "data": resting_hr_daily(days),
        "baseline": resting_hr_baseline(days),
    })


@app.get("/api/health/sleep")
def api_health_sleep():
    days = int(request.args.get("days", 30))
    return jsonify(sleep_daily(days))


@app.get("/api/health/spo2")
def api_health_spo2():
    days = int(request.args.get("days", 30))
    return jsonify(spo2_daily(days))


@app.get("/api/health/respiratory-rate")
def api_health_resp():
    days = int(request.args.get("days", 30))
    return jsonify(resp_rate_daily(days))


@app.get("/api/health/steps")
def api_health_steps():
    days = int(request.args.get("days", 30))
    return jsonify(steps_daily(days))


@app.get("/api/health/vo2")
def api_health_vo2():
    days = int(request.args.get("days", 90))
    return jsonify(vo2_trend(days))


@app.get("/api/health/nutrition")
def api_health_nutrition():
    days = int(request.args.get("days", 30))
    return jsonify(nutrition_daily(days))


@app.get("/api/health/workouts")
def api_health_workouts():
    days = int(request.args.get("days", 30))
    return jsonify(workout_sessions(days))


@app.post("/api/health/config")
def api_health_config():
    payload = request.get_json(force=True) or {}
    allowed = {"hr_max", "sleep_goal_hrs", "sleep_source_preference"}
    for k, v in payload.items():
        if k in allowed:
            set_config(k, str(v))
    return jsonify({"ok": True, "config": get_config()})


@app.post("/api/health/import/apple-health")
def api_health_import_apple():
    if "file" in request.files:
        f = request.files["file"]
        safe_name = re.sub(r"[^\w.\- ]", "_", f.filename or "health.json")
        dest = HEALTH_UPLOAD_DIR / safe_name
        f.save(str(dest))
        path = dest
    else:
        path = latest_health_json()
        if not path:
            return jsonify({"ok": False, "error": "No Health Auto Export file found in iCloud Drive"}), 404

    try:
        counts = parse_and_import(path)
        return jsonify({"ok": True, "counts": counts, "file": path.name})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.post("/api/health/import/macrofactor")
def api_health_import_mf():
    results = {}
    errors = {}

    nutr_file = request.files.get("nutrition")
    work_file = request.files.get("workouts")

    nutr_path = HEALTH_UPLOAD_DIR / "nutrition.csv" if nutr_file else NUTRITION_PATH
    work_path = HEALTH_UPLOAD_DIR / "workouts.csv" if work_file else WORKOUTS_PATH

    if nutr_file:
        nutr_file.save(str(nutr_path))
    if work_file:
        work_file.save(str(work_path))

    if nutr_path.exists():
        try:
            results["nutrition"] = _import_nutrition(nutr_path)
        except Exception as exc:
            errors["nutrition"] = str(exc)
    else:
        errors["nutrition"] = f"File not found: {nutr_path}"

    if work_path.exists():
        try:
            results["workouts"] = _import_workouts(work_path)
        except Exception as exc:
            errors["workouts"] = str(exc)
    else:
        errors["workouts"] = f"File not found: {work_path}"

    ok = not errors
    return jsonify({"ok": ok, "imported": results, "errors": errors}), (200 if ok else 207)


@app.get("/api/networth")
def api_networth():
    return jsonify(read_csv(NETWORTH_CSV))


@app.get("/api/holdings")
def api_holdings():
    return jsonify(read_csv(HOLDINGS_CSV))


@app.post("/api/networth/import")
def api_networth_import():
    payload = request.get_json(force=True) or {}
    try:
        super_aud = float(payload.get("super_aud", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "super_aud must be a number"}), 400
    try:
        count = import_networth_from_excel(super_aud)
    except FileNotFoundError as exc:
        return jsonify({"ok": False, "error": str(exc)}), 404
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": True, "imported": count})


def _get_gspread_client():
    try:
        import gspread  # noqa: PLC0415
        from google.oauth2.service_account import Credentials  # noqa: PLC0415
    except ImportError:
        raise RuntimeError("Install gspread and google-auth: pip install gspread google-auth")
    config = load_config()
    gs = config.get("google_sheets", {})
    creds_path = BASE_DIR / gs.get("credentials_path", "credentials.json")
    if not creds_path.exists():
        raise FileNotFoundError(f"Google credentials file not found: {creds_path}")
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(str(creds_path), scopes=scopes)
    return gspread.authorize(creds), gs


def import_networth_from_sheets() -> int:
    gc, gs = _get_gspread_client()
    sheet_id = gs.get("networth_sheet_id", "")
    if not sheet_id:
        raise ValueError("google_sheets.networth_sheet_id not set in config.json")

    tab = gs.get("networth_sheet_tab", "")
    spreadsheet = gc.open_by_key(sheet_id)
    ws = spreadsheet.worksheet(tab) if tab else spreadsheet.get_worksheet(0)
    records = ws.get_all_records()

    col_date = gs.get("networth_col_date", "Timestamp")
    col_everyday = gs.get("networth_col_everyday", "Everyday")
    col_savings = gs.get("networth_col_savings", "Savings")
    col_selfwealth = gs.get("networth_col_selfwealth", "SelfWealth")
    col_ibkr = gs.get("networth_col_ibkr", "IBKR")
    col_super = gs.get("networth_col_super", "")

    existing = {row["date"]: row for row in read_csv(NETWORTH_CSV)}
    imported = 0
    for record in records:
        raw_date = str(record.get(col_date, "")).strip()
        if not raw_date:
            continue
        try:
            date_str = parser.parse(raw_date, dayfirst=True).strftime("%Y-%m-%d")
        except Exception:
            continue
        everyday = float(record.get(col_everyday) or 0)
        savings = float(record.get(col_savings) or 0)
        selfwealth = float(record.get(col_selfwealth) or 0)
        ibkr = float(record.get(col_ibkr) or 0)
        super_aud = float(record.get(col_super) or 0) if col_super else 0.0
        cash = round(everyday + savings, 2)
        investments = round(selfwealth + ibkr, 2)
        existing[date_str] = {
            "date": date_str,
            "cash_aud": str(cash),
            "investments_aud": str(investments),
            "super_aud": str(round(super_aud, 2)),
            "total_aud": str(round(cash + investments + super_aud, 2)),
        }
        imported += 1

    write_csv(NETWORTH_CSV, NETWORTH_FIELDS, sorted(existing.values(), key=lambda r: r["date"]))
    return imported


def import_bills_from_sheets() -> int:
    gc, gs = _get_gspread_client()
    sheet_id = gs.get("bills_sheet_id", "")
    if not sheet_id:
        raise ValueError("google_sheets.bills_sheet_id not set in config.json")

    ws = gc.open_by_key(sheet_id).get_worksheet(0)
    records = ws.get_all_records()

    col_slug = gs.get("bills_col_slug", "Bill Type")
    col_amount = gs.get("bills_col_amount", "Amount")
    col_due_date = gs.get("bills_col_due_date", "Due Date")
    col_notes = gs.get("bills_col_notes", "Notes")

    bills = read_bills()
    added = 0
    for record in records:
        slug = str(record.get(col_slug, "")).strip().lower()
        if slug not in BILL_SLUGS:
            continue
        raw_due = str(record.get(col_due_date, "")).strip()
        try:
            due_date_str = parser.parse(raw_due).strftime("%Y-%m-%d")
        except Exception:
            continue
        if any(b.get("slug", "").lower() == slug and b.get("due_date") == due_date_str for b in bills):
            continue
        amount_raw = str(record.get(col_amount, "")).replace("$", "").replace(",", "").strip()
        amount = None
        try:
            if amount_raw:
                amount = float(amount_raw)
        except ValueError:
            pass
        notes = str(record.get(col_notes, "")).strip()
        append_bill({
            "slug": slug,
            "label": LABEL_DEFAULTS.get(slug, slug.title()),
            "total_amount": amount,
            "due_date": due_date_str,
            "recurrence": "monthly",
            "split_type": "fixed" if slug == "rent" else "equal",
            "notes": notes,
        })
        bills = read_bills()
        added += 1
    return added


@app.post("/api/networth/import-sheets")
def api_networth_import_sheets():
    try:
        count = import_networth_from_sheets()
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": True, "imported": count})


@app.post("/api/bills/sync-sheets")
def api_bills_sync_sheets():
    try:
        added = import_bills_from_sheets()
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500
    return jsonify({"ok": True, "added": added})


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5001)
